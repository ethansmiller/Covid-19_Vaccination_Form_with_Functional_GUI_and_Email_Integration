# -*- coding: utf-8 -*-
"""
Created on Sat Feb 20 21:11:22 2021
references include: https://www.codespeedy.com/how-to-add-days-to-date-in-python/
                    https://realpython.com/python-send-email/
                    https://www.educba.com/python-tkinter-label/
                    https://datatofish.com/entry-box-tkinter/
As well as sample code provided by the professor for sending mass emails and
manipulating/interacting with excel files
                    
@author: Ethan Miller, 011075077, EE104
"""

# openpyxl, tkinter, datetime, pandas, and smtplib imported
from openpyxl import *
import tkinter as tk
import datetime
import pandas as pd
import smtplib

#function used to check if the current date is three days before the second vaccination date
def date_check():
    
    first_date = first_vacc_field.get()
    first_date_temp = datetime.datetime.strptime(first_date, '%m/%d/%y')
    second_vacc_date = first_date_temp + datetime.timedelta(days=21)
    
    current_date = current_date_field.get()
    current_date_temp = datetime.datetime.strptime(current_date, '%m/%d/%y')
    check_date = second_vacc_date - datetime.timedelta(days=3)
    
    #if true, the function calls for a reminder to be sent
    if check_date == current_date_temp:
        reminder()
        
    #if false, the function passes and does not send the reminder  
    else:
        pass


#function used to send an email reminder, giving the patient a notice 3 days before their second appointment
def reminder():

    #login credentials for the secondary email used to email patient notifications
    your_name = "EE104 SJSU"
    your_email = "ee104test11@gmail.com"
    your_password = "EE104S2021"
    
    #accessing the gmail server via smtplib library
    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    server.ehlo()
    server.login(your_email, your_password)
    
    #defining the second vaccination date appointment based on the information provided
    #timedelta is used to add 21 days to the first appointment
    first_date = first_vacc_field.get()
    first_date_temp = datetime.datetime.strptime(first_date, '%m/%d/%y')
    second_vacc_date = first_date_temp + datetime.timedelta(days=21)

    #patient information is gathered to send an email
    name = first_name_field.get()
    email = email_field.get()
    first_vacc = first_vacc_field.get()
    second_vacc = second_vacc_date
    
    #the email sent as a reminder of the upcoming second appointment    
    full_email = ("From: {0} <{1}>\n"
                  "To: {2} <{3}>\n"
                  "Subject: 2nd Covid Vaccination Appointment reminder\n"
                  "Your first vaccination appointment was scheduled for {4}\n\n"
                  "This is a reminder that your second Covid vaccination appointment is scheduled for {5}\n\n"
                  "Sincerely,\nThe EE104 Medical Staff"
                  .format(your_name, your_email, name, email, first_vacc, second_vacc))

    #sending the email. a message is printed in the console for whether or not the email was sent successfully
    try:
        server.sendmail(your_email, [email], full_email)
        print('Email to {} successfully sent!\n\n'.format(email))
    except Exception as e:
        print('Email to {} could not be sent because {}\n\n'.format(email, str(e)))
   
    #the server is closed after the email is sent
    server.close()


#a function used to send the confirmation email after all information has been
#saved and the first and second vaccination appointments are set
#https://realpython.com/python-send-email/ was used as reference as well as
# sample code provided by the professor
def email():
    
    #login credentials for the secondary email used to email patient notifications
    your_name = "EE104 SJSU"
    your_email = "ee104test11@gmail.com"
    your_password = "EE104S2021"
    
    #accessing the gmail server via smtplib library
    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    server.ehlo()
    server.login(your_email, your_password)
    
    #defining the second vaccination date appointment based on the information provided
    #timedelta is used to add 21 days to the first appointment
    first_date = first_vacc_field.get()
    first_date_temp = datetime.datetime.strptime(first_date, '%m/%d/%y')
    second_vacc_date = first_date_temp + datetime.timedelta(days=21)
    
    #patient information is gathered to send an email
    name = first_name_field.get()
    email = email_field.get()
    first_vacc = first_vacc_field.get()
    second_vacc = second_vacc_date
    
    #the email sent as a confirmation of the two appointments
    full_email = ("From: {0} <{1}>\n"
                  "To: {2} <{3}>\n"
                  "Subject: 1st and 2nd Covid Vaccination Appointments Confirmation\n"
                  "Your first vaccination appointment is scheduled for {4}.\n\n"
                  "Your second Vaccination appointment will be for {5}\n\n"
                  "Sincerely,\nThe EE104 Medical Staff"
                  .format(your_name, your_email, name, email, first_vacc, second_vacc))
    
    #sending the email. a message is printed in the console for whether or not the email was sent successfully             
    try:
        server.sendmail(your_email, [email], full_email)
        print('Email to {} successfully sent!\n\n'.format(email))
    except Exception as e:
        print('Email to {} could not be sent because {}\n\n'.format(email, str(e)))
    
    #the server is closed after the email is sent
    server.close()
            

# opening the excel sheet for patient information and creating sheet object
path = '.\\Vacc_info.xlsx'
wb = load_workbook(path)
sheet = wb.active


#function used to create the layout of the excel sheet including column width and titles
def excel():
    
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 15
    
    sheet.cell(row=1, column=1).value = "First Name"
    sheet.cell(row=1, column=2).value = "Middle Name"
    sheet.cell(row=1, column=3).value = "Last Name"
    sheet.cell(row=1, column=4).value = "DoB"
    sheet.cell(row=1, column=5).value = "Phone Number"
    sheet.cell(row=1, column=6).value = "Email"
    sheet.cell(row=1, column=7).value = "First Vacc Date"
    sheet.cell(row=1, column=8).value = "Second Vacc Date"
    sheet.cell(row=1, column=9).value = "Current Date"


#function used to insert gathered info into the respective cells of the sheet
def insert():
    
    first_date = first_vacc_field.get()
    first_date_temp = datetime.datetime.strptime(first_date, '%m/%d/%y')
    second_vacc_date = first_date_temp + datetime.timedelta(days=21)
       
    current_row = sheet.max_row
    current_column = sheet.max_column
    
    sheet.cell(row=current_row + 1, column=1).value = first_name_field.get()
    sheet.cell(row=current_row + 1, column=2).value = middle_name_field.get()
    sheet.cell(row=current_row + 1, column=3).value = last_name_field.get()
    sheet.cell(row=current_row + 1, column=4).value = DoB_field.get()
    sheet.cell(row=current_row + 1, column=5).value = phone_field.get()
    sheet.cell(row=current_row + 1, column=6).value = email_field.get()
    sheet.cell(row=current_row + 1, column=7).value = first_vacc_field.get()
    sheet.cell(row=current_row + 1, column=8).value = second_vacc_date
    sheet.cell(row=current_row + 1, column=9).value = current_date_field.get()
    
    wb.save('.\\Vacc_info.xlsx')
    
       
#function used to create the popup window confirming that the information was saved
#and the confirmation/reminder is being sent via email
def confirmwindow():
    
    #popup window created
    popup = tk.Tk()
    popup.minsize(150,150)
    popup.wm_title('Confirmed')
    label = tk.Label(popup, width=100, height=10, text="A confirmation email has been sent, which will contain the second appointment date.\nAn email will also be sent 3 days prior to the second appointment.")
    label.pack()
    #button created that destroys the window when pressed
    okay = tk.Button(popup, text='okay', width=11, height=3, command= lambda: popup.destroy())
    okay.place(x=300, y=100)
    popup.mainloop()

#primary confirmation function that is called by the confirm button
#includes all previous functions that allow for the info gathered to be stored, and email to be sent
#the current date to be matched with the upcoming second appointment, and opening the confirmation
#window before the excel sheet is saved
def confirmation():
    
    insert()
    email()
    date_check()
    confirmwindow()
    excel()
    wb.save('.\\Vacc_info.xlsx')

#body of program containing the main root window and all widgets (buttons, labels, etc...)    
if __name__ == "__main__":

    root = tk.Tk()

   # creating the object that will hold the GUI of our medical info form

    root.title("Covid-19 Vaccination Form")
    
    excel()
    
    # text and entry were created with reference from https://datatofish.com/entry-box-tkinter/
    # and https://www.educba.com/python-tkinter-label/
    
    # the first canvas holds the entry box and labels for inputting the name of the patient
    canvas1 = tk.Canvas(root, width = 600, height = 220)
    canvas1.pack()
    
    toplabel = tk.Label(root, text= 'Patient information for Covid-19 Vaccination', font= 'helvetica')
    canvas1.create_window(300, 40, window=toplabel)
    
    current_date_field = tk.Entry(root, width=30)
    canvas1.create_window(35, 105, window=current_date_field)
    
    first_name_field = tk.Entry(root, width=25)
    canvas1.create_window(20, 180, window=first_name_field)
    
    middle_name_field = tk.Entry(root, width=25)
    canvas1.create_window(200, 180, window=middle_name_field)
    
    last_name_field = tk.Entry(root, width=25)
    canvas1.create_window(380, 180, window=last_name_field)
    
    label0 = tk.Label(root, text="Current Date:")
    canvas1.create_window(-21, 80, window=label0)
    
    label_date = tk.Label(root, text="MM/DD/YY")
    canvas1.create_window(95, 80, window=label_date)
    
    label1 = tk.Label(root, text= 'First Name:')
    canvas1.create_window(-26,156,window= label1)
    
    label2 = tk.Label(root, text= 'Middle Name:')
    canvas1.create_window(161, 156, window=label2)
    
    label3 = tk.Label(root, text= 'Last Name:')
    canvas1.create_window(333, 156, window=label3)

    #the second canvas holds the entry boxes and labels for inputting the date of birth

    canvas2 = tk.Canvas(root, width = 800, height = 100)
    canvas2.pack()
    
    DoB_field = tk.Entry(root, width=30)
    canvas2.create_window(135, 45, window=DoB_field)
      
    label4 = tk.Label(root, text= 'Date of Birth:')
    canvas2.create_window(78, 21, window=label4)
    
    label5 = tk.Label(root, text= 'MM/DD/YYYY')
    canvas2.create_window(190, 21, window=label5)
    
    #the third canvas holds the entry boxes and labels for inputting phone number and email

    canvas3 = tk.Canvas(root, width = 800, height = 450)
    canvas3.pack()
    
    titlelabel1 = tk.Label(root, text= 'Contact Information:', font= 'bold')
    canvas3.create_window(112, 10, window=titlelabel1)
    
    phone_field = tk.Entry(root, width=35)
    canvas3.create_window(150, 80, window=phone_field)
    
    label6 = tk.Label(root, text= '(XXX) - XXX - XXXX')
    canvas3.create_window(190, 55, window=label6)
    
    email_field = tk.Entry(root, width=40)
    canvas3.create_window(165, 160, window=email_field)
    
    label7 = tk.Label(root, text= 'Phone Number:')
    canvas3.create_window(86, 56, window=label7)
    
    label8 = tk.Label(root, text= 'Email Address:')
    canvas3.create_window(83, 136, window=label8)
    
    #Also in the third canvas is the final input for the vaccination date
    
    titlelabel2 = tk.Label(root, text= 'Vaccination Information:', font= 'bold')
    canvas3.create_window(127, 235, window=titlelabel2)
    
    first_vacc_field = tk.Entry(root, width=30)
    canvas3.create_window(135, 310, window=first_vacc_field)      
    
    label9 = tk.Label(root, text= 'First Vaccination Date:')
    canvas3.create_window(102, 286, window=label9)
    
    label10 = tk.Label(root, text = 'MM/DD/YY')
    canvas3.create_window(200, 286, window=label10)
    
    label11 = tk.Label(root, text= '(Your second Vaccination Date will be sent via a confirmation email)')
    canvas3.create_window(410, 340, window=label11)
    
    #button to confirm information. confirmation function is called, which was described above
    
    Confirm1 = tk.Button(root, text='confirm', width=11, height=3, command= lambda: 
    confirmation())
    canvas3.create_window(400, 400, window=Confirm1)
    
    excel()

    
    root.mainloop()